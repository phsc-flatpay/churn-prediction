# Fabric notebook source

# METADATA ********************

# META {
# META   "kernel_info": {
# META     "name": "synapse_pyspark"
# META   },
# META   "dependencies": {
# META     "lakehouse": {
# META       "default_lakehouse": "54c8eb17-9ab5-4b14-9a01-728630a243fb",
# META       "default_lakehouse_name": "ml_ops",
# META       "default_lakehouse_workspace_id": "e03dbe51-4f30-4e31-a84f-647f6b831f58",
# META       "known_lakehouses": [
# META         {
# META           "id": "54c8eb17-9ab5-4b14-9a01-728630a243fb"
# META         }
# META       ]
# META     }
# META   }
# META }

# CELL ********************

from datetime import datetime
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Configuration
TIERS = ["high", "medium", "low"]                       # tiers to export

DELTA_ROOT = (
    "abfss://Churn@onelake.dfs.fabric.microsoft.com/"
    "ml_ops.Lakehouse/Tables"
)

TMP_DIR = "/tmp/exports"                                # drivelocal temp dir
LAKE_FOLDER = (
    "abfss://Churn@onelake.dfs.fabric.microsoft.com/"
    "ml_ops.Lakehouse/Files/exports"
)

os.makedirs(TMP_DIR, exist_ok=True)                     # ensure tmp dir exists

# Export loop
export_paths = {}                                       # collect final URIs

for tier in TIERS:
    # Paths
    delta_path = f"{DELTA_ROOT}/Churn_Predictions_{tier}_TPV"
    tmp_excel  = f"{TMP_DIR}/Churn_Predictions_{tier}_TPV.xlsx"
    lake_excel = f"{LAKE_FOLDER}/Churn_Predictions_{tier}_TPV.xlsx"

    print(f" Reading  : {delta_path}")
    print(f" Writing  : {tmp_excel}")

    # Load Delta to Pandas
    df_spark = spark.read.format("delta").load(delta_path)
    df_pd    = df_spark.toPandas()

    # Write local Excel
    df_pd.to_excel(tmp_excel, index=False, engine="openpyxl")
    print("   Local Excel written.")


    # Auto-fit columns
    wb = load_workbook(tmp_excel)
    ws = wb.active                # first sheet
    for idx, col in enumerate(df_pd.columns, 1):
        max_len = max(
            df_pd[col].astype(str).map(len).max(),  # longest cell
            len(col)                                # header
        )
        ws.column_dimensions[get_column_letter(idx)].width = max_len + 4  # +4 padding
    wb.save(tmp_excel)


    # Copy to lakehouse
    try:
        from notebookutils import mssparkutils

        mssparkutils.fs.mkdirs(LAKE_FOLDER)             # idempotent directory create
        mssparkutils.fs.cp(f"file:{tmp_excel}", lake_excel, True)
        print(f"   Copied to lakehouse : {lake_excel}")

        export_paths[tier] = lake_excel

    except ImportError:
        # Running outside Fabric fall back to plain print
        print(f" ! notebookutils unavailable – cannot copy {tier} file.")
        export_paths[tier] = lake_excel  # still record intended target

#
# Emit shareable URLs (for downstream steps)
#
base_url = "https://onelake.dfs.fabric.microsoft.com"
shareable = {
    t: f"{base_url}/{v.replace('abfss://','').replace('@onelake.dfs.fabric.microsoft.com','')}"
    for t, v in export_paths.items()
}

try:
    mssparkutils.notebook.exit(json.dumps(shareable))
except Exception:
    # Fallback for local debugging
    print(json.dumps(shareable, indent=2))


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

from notebookutils import mssparkutils

lake_folder = "abfss://Churn@onelake.dfs.fabric.microsoft.com/ml_ops.Lakehouse/Files/exports"

# all items under the abfs folder
files = mssparkutils.fs.ls(lake_folder)

# print each filename and size
for f in files:
    print(f.name, f.size)


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }
